#!/usr/bin/env python3
"""把公开项目 Excel 转换成静态网页使用的 JSON。"""

from __future__ import annotations

import argparse
import hashlib
import json
import posixpath
import re
import shutil
import struct
import sys
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import unquote
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook


FIELD_ALIASES = {
    "id": ("项目编号", "编号", "序号"),
    "title": ("项目名称", "设计名称", "课题名称", "名称"),
    "series": ("项目系列", "系列"),
    "mcuFamily": ("单片机分类", "单片机系列", "主控分类"),
    "mcuModel": ("单片机型号", "主控型号", "型号"),
    "usages": ("项目用途", "用途", "功能分类"),
    "modules": ("使用模块", "模块", "硬件模块"),
    "simulationImage": ("仿真图片", "仿真图"),
    "hardwareImage": ("实物图片", "实物图"),
    "description": ("项目简介", "简介", "功能简介"),
    "keywords": ("搜索关键词", "关键词", "标签"),
    "visible": ("是否展示", "展示", "上架"),
    "sort": ("排序", "显示顺序"),
    "simulationPrice": ("仿真+仿真代码", "仿真及代码", "仿真价格"),
    "pcbPrice": ("原理图+PCB设计", "原理图及PCB", "PCB价格"),
    "hardwarePrice": ("硬件实物+配套硬件代码", "硬件实物及代码", "硬件价格"),
    "thesisPrice": ("论文", "论文价格"),
}

FORBIDDEN_HEADERS = {
    "下载链接",
    "下载地址",
    "资料介绍链接",
    "资料链接",
    "资料主要内容",
    "资料内容",
}
FALSE_VALUES = {"否", "不展示", "下架", "false", "0", "no"}
ABSENT_PRICE_VALUES = {"无"}
PRICE_FIELDS = (
    ("仿真+仿真代码", "simulationPrice"),
    ("原理图+PCB设计", "pcbPrice"),
    ("硬件实物+配套硬件代码", "hardwarePrice"),
    ("论文", "thesisPrice"),
)
IMAGE_FIELDS = (
    ("仿真图片", "simulationImage"),
    ("实物图片", "hardwareImage"),
)
DISPIMG_RE = re.compile(
    r'^=?\s*(?:_xlfn\.)?DISPIMG\(\s*"([^"]+)"\s*,\s*[01]\s*\)\s*$',
    re.IGNORECASE,
)
WPS_NAMESPACES = {
    "etc": "http://www.wps.cn/officeDocument/2017/etCustomData",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}
RELATIONSHIP_NAMESPACE = "{http://schemas.openxmlformats.org/package/2006/relationships}"
RELATIONSHIP_IMAGE_TYPE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image"
)
RELATIONSHIP_EMBED = f"{{{WPS_NAMESPACES['r']}}}embed"
MAX_IMAGE_BYTES = 20 * 1024 * 1024
MAX_TOTAL_IMAGE_BYTES = 500 * 1024 * 1024
PROJECT_IMAGE_URL = "./assets/generated/projects"
REPOSITORY_ROOT = Path(__file__).resolve().parents[1]
PROJECT_IMAGE_DIRECTORY = REPOSITORY_ROOT / "site" / "assets" / "generated" / "projects"

USAGE_RULES = (
    ("温湿度检测", r"温湿度|温度|湿度|测温|体温"),
    ("超声波测距", r"超声波|测距|距离"),
    ("环境监测", r"环境|空气质量|烟雾|气体|粉尘|PM\s*2\.5|甲醛|光照"),
    ("安防报警", r"报警|安防|火焰|防盗|门禁|入侵"),
    ("智能家居", r"智能家居|台灯|窗帘|门锁|照明|风扇|晾衣"),
    ("农业控制", r"大棚|灌溉|土壤|农业|养殖|花卉"),
    ("医疗健康", r"心率|血压|病床|输液|健康|脉搏|血氧"),
    ("智能交通", r"小车|循迹|交通灯|车位|停车|红绿灯|车辆"),
    ("显示与交互", r"点阵|LCD|OLED|显示|广告牌|触摸"),
    ("物联网", r"物联网|WiFi|云平台|远程|上位机"),
    ("电机控制", r"电机|舵机|步进|PWM"),
    ("测量仪器", r"电子秤|频率|信号发生器|电压|电流|测量|计数器"),
    ("音频语音", r"语音|音响|音乐|MP3|播报"),
)

MODULE_RULES = (
    ("DHT11温湿度传感器", r"DHT11"),
    ("DS18B20温度传感器", r"DS18B20"),
    ("温湿度传感器", r"温湿度"),
    ("HC-SR04超声波模块", r"HC-?SR04|超声波"),
    ("ESP8266 WiFi模块", r"ESP8266"),
    ("WiFi通信模块", r"WiFi|物联网|云平台"),
    ("蓝牙模块", r"蓝牙|Bluetooth"),
    ("OLED显示屏", r"OLED"),
    ("LCD1602显示屏", r"LCD1602"),
    ("LCD显示屏", r"LCD|液晶"),
    ("LED点阵屏", r"点阵|LED屏"),
    ("RFID模块", r"RFID|射频卡|刷卡"),
    ("GPS定位模块", r"GPS|定位"),
    ("GSM通信模块", r"GSM|SIM800|短信"),
    ("语音识别模块", r"语音识别|语音控制"),
    ("烟雾传感器", r"烟雾"),
    ("火焰传感器", r"火焰"),
    ("土壤湿度传感器", r"土壤湿度"),
    ("继电器模块", r"继电器"),
    ("电机驱动模块", r"电机|步进"),
    ("舵机", r"舵机"),
    ("蜂鸣器", r"蜂鸣器|报警"),
)


class ValidationError(Exception):
    """Excel 内容不符合发布要求。"""


def value_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def split_list(value: object) -> list[str]:
    values = [item.strip() for item in re.split(r"[、,，;；|]+", value_to_text(value))]
    return list(dict.fromkeys(item for item in values if item))


def infer_by_rules(text: str, rules: tuple[tuple[str, str], ...]) -> list[str]:
    return [name for name, pattern in rules if re.search(pattern, text, re.IGNORECASE)]


def infer_mcu(text: str) -> str:
    checks = (
        ("STM32", r"STM32"),
        ("ESP32", r"ESP32"),
        ("Arduino", r"Arduino"),
        ("MSP430", r"MSP430"),
        ("PIC", r"PIC\d*"),
        ("AVR", r"AVR|ATmega"),
        ("51单片机", r"STC|AT89|51单片机|\b51\b"),
    )
    for name, pattern in checks:
        if re.search(pattern, text, re.IGNORECASE):
            return name
    return ""


def parse_sort(value: object, row_number: int) -> int:
    text = value_to_text(value)
    if not text:
        return row_number - 1
    try:
        number = float(text)
    except ValueError as error:
        raise ValidationError(f"第{row_number}行：排序必须是数字") from error
    if not number.is_integer():
        raise ValidationError(f"第{row_number}行：排序必须是整数")
    return int(number)


def parse_price(value: object, row_number: int, label: str) -> int | float | str | None:
    text = value_to_text(value)
    if value is None or text == "" or text in ABSENT_PRICE_VALUES:
        return None
    if isinstance(value, bool):
        raise ValidationError(f"第{row_number}行：{label}必须填写金额、面议、无或留空")
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        if text in {"面议", "咨询", "价格咨询"}:
            return text
        cleaned = re.sub(r"[¥￥,，\s]", "", text)
        try:
            number = float(cleaned)
        except ValueError as error:
            raise ValidationError(
                f"第{row_number}行：{label}必须填写金额、面议、无或留空"
            ) from error
    if number < 0:
        raise ValidationError(f"第{row_number}行：{label}不能为负数")
    return int(number) if number.is_integer() else round(number, 2)


def image_info(data: bytes) -> tuple[str, int, int]:
    if (
        data.startswith(b"\x89PNG\r\n\x1a\n")
        and len(data) >= 24
        and data[12:16] == b"IHDR"
    ):
        width, height = struct.unpack(">II", data[16:24])
        if width and height:
            return "png", width, height

    if data.startswith(b"\xff\xd8"):
        position = 2
        start_of_frame = {
            0xC0, 0xC1, 0xC2, 0xC3, 0xC5, 0xC6, 0xC7,
            0xC9, 0xCA, 0xCB, 0xCD, 0xCE, 0xCF,
        }
        while position + 4 <= len(data):
            while position < len(data) and data[position] == 0xFF:
                position += 1
            if position >= len(data):
                break
            marker = data[position]
            position += 1
            if marker in {0x01, *range(0xD0, 0xDA)}:
                continue
            if position + 2 > len(data):
                break
            length = int.from_bytes(data[position:position + 2], "big")
            if length < 2 or position + length > len(data):
                break
            if marker in start_of_frame and length >= 7:
                height = int.from_bytes(data[position + 3:position + 5], "big")
                width = int.from_bytes(data[position + 5:position + 7], "big")
                if width and height:
                    return "jpg", width, height
            position += length

    raise ValidationError("项目图片仅支持 JPG 或 PNG 格式")


def safe_media_member(target: str, members: set[str]) -> str:
    decoded = unquote(target.replace("\\", "/"))
    if not decoded or "\x00" in decoded or decoded.startswith("/"):
        raise ValidationError("WPS图片路径无效")
    if ".." in decoded.split("/"):
        raise ValidationError("WPS图片路径不能包含上级目录")
    member = posixpath.normpath(posixpath.join("xl", decoded))
    if not member.startswith("xl/media/") or member not in members:
        raise ValidationError("WPS图片媒体文件缺失")
    return member


def load_wps_images(input_path: Path) -> dict[str, dict]:
    with ZipFile(input_path) as archive:
        members = archive.namelist()
        if len(members) != len(set(members)):
            raise ValidationError("Excel压缩包包含重复文件")
        member_set = set(members)
        cell_images_path = "xl/cellimages.xml"
        relationships_path = "xl/_rels/cellimages.xml.rels"
        present = {cell_images_path, relationships_path}.intersection(member_set)
        if not present:
            return {}
        if len(present) != 2:
            raise ValidationError("WPS图片索引不完整")

        relationships = {}
        relationship_root = ElementTree.fromstring(archive.read(relationships_path))
        for relationship in relationship_root.findall(
            f"{RELATIONSHIP_NAMESPACE}Relationship"
        ):
            if relationship.get("TargetMode") == "External":
                raise ValidationError("项目图片不能使用外部链接")
            if relationship.get("Type") == RELATIONSHIP_IMAGE_TYPE:
                relationships[relationship.get("Id", "")] = relationship.get("Target", "")

        images = {}
        total_bytes = 0
        image_root = ElementTree.fromstring(archive.read(cell_images_path))
        for item in image_root.findall("etc:cellImage", WPS_NAMESPACES):
            properties = item.find(".//xdr:cNvPr", WPS_NAMESPACES)
            blip = item.find(".//a:blip", WPS_NAMESPACES)
            if properties is None or blip is None:
                raise ValidationError("WPS图片节点不完整")
            image_id = properties.get("name", "")
            relationship_id = blip.get(RELATIONSHIP_EMBED, "")
            target = relationships.get(relationship_id)
            if not image_id or not target:
                raise ValidationError("WPS图片ID或关系缺失")
            if image_id in images:
                raise ValidationError(f"WPS图片ID重复：{image_id}")

            member = safe_media_member(target, member_set)
            image_bytes = archive.getinfo(member).file_size
            if image_bytes > MAX_IMAGE_BYTES:
                raise ValidationError(f"图片“{image_id}”超过20MB")
            total_bytes += image_bytes
            if total_bytes > MAX_TOTAL_IMAGE_BYTES:
                raise ValidationError("Excel内图片总大小超过500MB")
            data = archive.read(member)
            extension, width, height = image_info(data)
            images[image_id] = {
                "data": data,
                "extension": extension,
                "width": width,
                "height": height,
            }
        return images


def parse_image(
    value: object,
    row_number: int,
    label: str,
    embedded_images: dict[str, dict],
    assets: dict[str, bytes] | None,
) -> dict | None:
    text = value_to_text(value)
    if not text:
        return None
    match = DISPIMG_RE.fullmatch(text)
    if not match:
        raise ValidationError(f"第{row_number}行：{label}必须使用WPS嵌入单元格图片或留空")
    image_id = match.group(1)
    image = embedded_images.get(image_id)
    if image is None:
        raise ValidationError(f"第{row_number}行：{label}对应的图片文件缺失")
    filename = f"{hashlib.sha256(image['data']).hexdigest()[:20]}.{image['extension']}"
    if assets is not None:
        assets[filename] = image["data"]
    return {
        "src": f"{PROJECT_IMAGE_URL}/{filename}",
        "width": image["width"],
        "height": image["height"],
    }


def resolve_columns(headers: list[str]) -> dict[str, int]:
    duplicate_headers = sorted({header for header in headers if headers.count(header) > 1 and header})
    if duplicate_headers:
        raise ValidationError(f"表头重复：{'、'.join(duplicate_headers)}")

    forbidden = sorted(FORBIDDEN_HEADERS.intersection(headers))
    if forbidden:
        raise ValidationError(
            f"公开Excel不能包含以下字段：{'、'.join(forbidden)}。请删除后再上传。"
        )

    columns: dict[str, int] = {}
    for canonical, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            if alias in headers:
                columns[canonical] = headers.index(alias)
                break

    missing = [name for name in ("id", "title", "mcuFamily") if name not in columns]
    if missing:
        labels = {"id": "项目编号", "title": "项目名称", "mcuFamily": "单片机分类"}
        raise ValidationError(f"缺少必填表头：{'、'.join(labels[name] for name in missing)}")
    return columns


def read_workbook_rows(input_path: Path, data_only: bool) -> list[tuple]:
    workbook = load_workbook(input_path, read_only=True, data_only=data_only)
    try:
        preferred = next(
            (name for name in ("项目链接清单", "项目数据库") if name in workbook.sheetnames),
            None,
        )
        sheet = workbook[preferred] if preferred else workbook.active
        return list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()


def parse_workbook(input_path: Path, assets: dict[str, bytes] | None = None) -> dict:
    all_rows = read_workbook_rows(input_path, data_only=True)
    formula_rows = read_workbook_rows(input_path, data_only=False)
    rows = iter(all_rows)

    try:
        headers = [value_to_text(value) for value in next(rows)]
    except StopIteration as error:
        raise ValidationError("Excel没有任何内容") from error

    columns = resolve_columns(headers)
    data_rows = list(rows)
    formula_data_rows = formula_rows[1:]
    projects: list[dict] = []
    seen_ids: dict[str, int] = {}

    def get(row: tuple, field: str) -> object:
        index = columns.get(field)
        return row[index] if index is not None and index < len(row) else None

    def get_image(row_index: int, row: tuple, field: str) -> object:
        formula_row = formula_data_rows[row_index] if row_index < len(formula_data_rows) else ()
        formula_value = get(formula_row, field)
        return formula_value if value_to_text(formula_value) else get(row, field)

    has_project_images = any(
        value_to_text(get(row, "visible")).lower() not in FALSE_VALUES
        and any(
            value_to_text(get_image(row_index, row, field))
            for _label, field in IMAGE_FIELDS
        )
        for row_index, row in enumerate(data_rows)
    )
    embedded_images = load_wps_images(input_path) if has_project_images else {}

    for row_index, row in enumerate(data_rows):
        row_number = row_index + 2
        if not any(value_to_text(value) for value in row):
            continue
        if value_to_text(get(row, "visible")).lower() in FALSE_VALUES:
            continue

        project_code = value_to_text(get(row, "id"))
        title = value_to_text(get(row, "title"))
        series = value_to_text(get(row, "series"))
        mcu_family = value_to_text(get(row, "mcuFamily")) or infer_mcu(title)
        errors = []
        if not project_code:
            errors.append("项目编号为空")
        if not title:
            errors.append("项目名称为空")
        if not mcu_family:
            errors.append("单片机分类为空且无法从名称识别")
        if errors:
            raise ValidationError(f"第{row_number}行：{'；'.join(errors)}")

        project_id = f"{series}::{project_code}" if series else project_code
        key = project_id.casefold()
        if key in seen_ids:
            raise ValidationError(
                f"第{row_number}行：项目系列和编号“{series} / {project_code}”与第{seen_ids[key]}行重复"
            )
        seen_ids[key] = row_number

        usages = split_list(get(row, "usages")) or infer_by_rules(title, USAGE_RULES) or ["综合应用"]
        modules = split_list(get(row, "modules")) or infer_by_rules(title, MODULE_RULES)
        prices = []
        for label, field in PRICE_FIELDS:
            parsed_price = parse_price(get(row, field), row_number, label)
            if parsed_price is not None:
                prices.append({"label": label, "price": parsed_price})
        project = {
            "id": project_id,
            "code": project_code,
            "title": title,
            "series": series,
            "mcuFamily": mcu_family,
            "mcuModel": value_to_text(get(row, "mcuModel")),
            "usages": usages,
            "modules": modules,
            "description": value_to_text(get(row, "description")),
            "keywords": split_list(get(row, "keywords")),
            "prices": prices,
            "sort": parse_sort(get(row, "sort"), row_number),
        }
        for label, field in IMAGE_FIELDS:
            image = parse_image(
                get_image(row_index, row, field),
                row_number,
                label,
                embedded_images,
                assets,
            )
            if image is not None:
                project[field] = image
        projects.append(project)

    if not projects:
        raise ValidationError("没有可展示的项目，请检查“是否展示”列和必填字段")

    projects.sort(key=lambda project: (project["sort"], project["id"].casefold()))
    return {
        "meta": {
            "source": input_path.name,
            "generatedAt": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
            "count": len(projects),
        },
        "projects": projects,
    }


def write_json(payload: dict, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary = output_path.with_suffix(f"{output_path.suffix}.tmp")
    temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    temporary.replace(output_path)


def write_assets(
    assets: dict[str, bytes], repository_root: Path = REPOSITORY_ROOT
) -> None:
    root = repository_root.resolve()
    output_path = repository_root / "site" / "assets" / "generated" / "projects"
    if output_path.resolve() != root / "site" / "assets" / "generated" / "projects":
        raise ValidationError("图片输出目录必须位于项目的固定目录内")
    current = repository_root
    for part in ("site", "assets", "generated", "projects"):
        current /= part
        if current.exists() and current.is_symlink():
            raise ValidationError("图片输出目录不能经过符号链接")
    if output_path.exists():
        if not output_path.is_dir():
            raise ValidationError("图片输出路径不是文件夹")
        shutil.rmtree(output_path)
    output_path.mkdir(parents=True, exist_ok=True)
    for filename, data in assets.items():
        (output_path / filename).write_bytes(data)


def main() -> int:
    parser = argparse.ArgumentParser(description="解析项目数据库Excel并生成网页数据")
    parser.add_argument("--input", type=Path, default=Path("data/项目链接清单.xlsx"))
    parser.add_argument("--output", type=Path, default=Path("site/data/projects.json"))
    parser.add_argument("--check-only", action="store_true")
    args = parser.parse_args()

    try:
        assets = None if args.check_only else {}
        payload = parse_workbook(args.input, assets)
        if not args.check_only:
            write_assets(assets)
            write_json(payload, args.output)
        print(f"Excel解析成功：{len(payload['projects'])}个项目")
        return 0
    except (BadZipFile, ElementTree.ParseError, OSError, ValidationError) as error:
        print(f"Excel校验失败：{error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
