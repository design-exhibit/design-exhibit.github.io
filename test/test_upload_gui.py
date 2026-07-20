import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "上传工具"))

from upload_gui import (  # noqa: E402
    DATA_PATH,
    UploadError,
    is_expected_remote,
    upload_workbook,
    validate_workbook,
    validate_workbook_with_visibility_summary,
    workbook_summary,
)


class UploadGuiBackendTest(unittest.TestCase):
    def test_validate_current_workbook(self):
        payload, digest = validate_workbook(DATA_PATH)
        summary = workbook_summary(payload)

        projects = payload["projects"]
        expected_images = sum(
            bool(project.get("simulationImage")) + bool(project.get("hardwareImage"))
            for project in projects
        )
        self.assertGreater(summary["projects"], 0)
        self.assertEqual(summary["projects"], len(projects))
        self.assertEqual(summary["images"], expected_images)
        self.assertEqual(len(digest), 64)

    def test_visibility_summary_lists_hidden_rows_without_affecting_projects(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "项目链接清单.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "项目链接清单"
            sheet.append(["项目编号", "项目名称", "单片机分类", "是否展示"])
            sheet.append(["P001", "展示项目", "STM32", "是"])
            sheet.append(["P002", "下架项目", "STM32", "否"])
            sheet.append(["  ", None, None, None])
            sheet.append(["P003", "不展示项目", "STM32", "下架"])
            workbook.save(path)

            payload, digest, summary = validate_workbook_with_visibility_summary(path)

        self.assertEqual(len(digest), 64)
        self.assertEqual(summary["dataRows"], 4)
        self.assertEqual(summary["shown"], 1)
        self.assertEqual(len(payload["projects"]), 1)
        self.assertEqual(summary["blankRows"], [4])
        self.assertEqual(
            summary["notDisplayed"],
            [
                {"row": 3, "code": "P002", "visible": "否"},
                {"row": 5, "code": "P003", "visible": "下架"},
            ],
        )

    def test_reject_changed_file_before_git(self):
        with self.assertRaisesRegex(UploadError, "检测后发生了变化"):
            upload_workbook(DATA_PATH, "0" * 64, lambda _message: None)

    def test_accept_expected_ssh_and_https_remote(self):
        self.assertTrue(
            is_expected_remote(
                "git@github.com:design-exhibit/design-exhibit.github.io.git"
            )
        )
        self.assertTrue(
            is_expected_remote(
                "https://github.com/design-exhibit/design-exhibit.github.io.git"
            )
        )
        self.assertFalse(is_expected_remote("git@github.com:other/repository.git"))
        self.assertFalse(
            is_expected_remote(
                "https://evilgithub.com/design-exhibit/design-exhibit.github.io.git"
            )
        )
        self.assertFalse(
            is_expected_remote(
                "git@evilgithub.com:design-exhibit/design-exhibit.github.io.git"
            )
        )


if __name__ == "__main__":
    unittest.main()
